#!/usr/bin/env python
# -*- coding: utf-8 -*-
# vim: ai ts=4 sts=4 et sw=4 nu

from __future__ import (unicode_literals, absolute_import,
                        division, print_function)
import re
import openpyxl

from datetime import datetime
# from optparse import make_option
from django.conf import settings
from django.core.management.base import BaseCommand

from desk.models import Entity, EntityType
from repatriate.models import (Person, targets)


# def str_to_date(str_date):
#     print (type(str_date))
#     return datetime.strptime(str_date, "%m/%d/%y")


def str_to_date(str_date):
    return datetime.strptime(str_date, "%d/%m/%Y")


class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument(
            '-f',
            help='CSV file to import from',
            action='store',
            dest='input_file'
        )

    def handle(self, *args, **options):
        self.read_xlsx_file(options.get('input_file'))

    def read_xlsx_file(self, file_source):
        print("DDD")
        book = openpyxl.load_workbook(file_source)
        print("JFJF")
        for sheetname in book.sheetnames:
            sheet = book[sheetname]
            print(sheet, "-" * 50)
            r = sheet.max_row
            for i in range(13, r + 1):
                c = 1
                print(sheet.cell(row=i, column=c).value)
                # locality, ok = Entity.objects.get_or_create(
                #     slug="{}_cercle".format(cercle.lower()), defaults={
                #         "name": cercle.upper(),
                #         "type": EntityType.objects.get(slug="cercle"),
                #         "parent": Entity.objects.get(slug="{}_region".format(region.lower()))})
                # number_social_name = sheet.cell(row=i, column=c).value

                # c += 1
                # activity_domaine = sheet.cell(row=i, column=c).value
                # c += 1
                # spinneret = sheet.cell(row=i, column=c).value
                # c += 1
                # head_office_vq = sheet.cell(row=i, column=c).value
                # c += 1
                # head_office_c = sheet.cell(row=i, column=c).value
                # c += 1
                # registration_number = sheet.cell(row=i, column=c).value
                # c += 1
                # registration_date = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_dadhesrents_m = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_dadhesrents_f = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_gestion_comite_m = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_gestion_comite_f = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_commission_surveillance_m = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_commission_surveillance_f = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_conseil_administration_m = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_conseil_administration_f = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_conseil_surveillance_m = sheet.cell(row=i, column=c).value
                # c += 1
                # nbre_conseil_surveillance_f = sheet.cell(row=i, column=c).value
                # c += 1
                # social_intial_capital = sheet.cell(row=i, column=c).value
                # c += 1
                # nb_emplois_crees = sheet.cell(row=i, column=c).value
                # data = {
                #     "locality": locality,
                #     "number_social_name": number_social_name,
                #     "category": category,
                #     "activity_domaine": activity_domaine.strip().upper(),
                #     "spinneret": "" if not spinneret else spinneret.strip().upper(),
                #     "head_office_vq": head_office_vq,
                #     "head_office_c": head_office_c,
                #     "registration_date": registration_date,
                #     "nbre_dadhesrents_m": nbre_dadhesrents_m,
                #     "nbre_dadhesrents_f": nbre_dadhesrents_f,
                #     "nbre_gestion_comite_m": nbre_gestion_comite_m,
                #     "nbre_gestion_comite_f": nbre_gestion_comite_f,
                #     "nbre_commission_surveillance_m": nbre_commission_surveillance_m,
                #     "nbre_commission_surveillance_f": nbre_commission_surveillance_f,
                #     "nbre_conseil_administration_m": nbre_conseil_administration_m,
                #     "nbre_conseil_administration_f": nbre_conseil_administration_f,
                #     "nbre_conseil_surveillance_m": nbre_conseil_surveillance_m,
                #     "nbre_conseil_surveillance_f": nbre_conseil_surveillance_f,
                #     "social_intial_capital": social_intial_capital,
                #     "nb_emplois_crees": nb_emplois_crees,
                # }
                # cc, ok = CooperativeCompanies.objects.update_or_create(
                #     registration_number=registration_number, defaults=data)
                # # print(rep_scoop.locality, rep_scoop.social_intial_capital, rep_scoop.nb_emplois_crees)
                # print(cc)
